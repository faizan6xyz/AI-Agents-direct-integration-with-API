import os
import re
import csv
import hmac
import hashlib
import logging
import threading
from datetime import datetime, timezone
from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook
VERIFY_TOKEN = os.environ.get("WHATSAPP_VERIFY_TOKEN")
APP_SECRET = os.environ.get("WHATSAPP_APP_SECRET")  # Meta App Dashboard > Settings > Basic
CSV_FILE = r"Analytics/Report/whatsapp_messages.csv"
EXCEL_FILE = r"Analytics/Report/whatsapp_messages.xlsx"
COLUMNS_NAME = ["Timestamp", "Sender Number", "Message"]
MAX_MESSAGE_LENGTH = 4000  # guards against pathological/huge payloads bloating the log
file_lock = threading.Lock()
app = Flask(__name__)
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
log = logging.getLogger("whatsapp_logger")
if not VERIFY_TOKEN or not APP_SECRET:
    log.warning("WHATSAPP_VERIFY_TOKEN and/or WHATSAPP_APP_SECRET are not set. "
        "Set them as environment variables before running in production.")

def is_valid_signature(req) -> bool:
    if not APP_SECRET:
        log.error("APP_SECRET not configured — refusing to process webhook.")
        return False
    signature_header = req.headers.get("X-Hub-Signature-256", "")
    if not signature_header.startswith("sha256="):
        log.warning("Missing or malformed X-Hub-Signature-256 header.")
        return False
    received_sig = signature_header.split("sha256=", 1)[1]
    expected_sig = hmac.new(
        APP_SECRET.encode("utf-8"), req.get_data(), hashlib.sha256
    ).hexdigest()
    return hmac.compare_digest(received_sig, expected_sig)

def sanitize_field(value: str) -> str:
    if value is None:
        return ""
    value = str(value)
    value = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", value)  # control chars, keep \n/\t out too if you prefer
    value = value.replace("\x00", "")
    if value[:1] in ("=", "+", "-", "@"):
        value = "'" + value
    return value[:MAX_MESSAGE_LENGTH]

def sanitize_sender(value: str) -> str:
    if not value:
        return "unknown"
    value = re.sub(r"[^0-9+]", "", str(value))
    return value[:20] or "unknown"

def ensure_csv_exists():
    if not os.path.exists(CSV_FILE):
        with open(CSV_FILE, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(COLUMNS_NAME)

def ensure_excel_exists():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Messages"
        ws.append(COLUMNS_NAME)
        wb.save(EXCEL_FILE)

def log_message(timestamp: str, sender: str, message: str):
    sender = sanitize_sender(sender)
    message = sanitize_field(message)
    with file_lock:
        try:
            ensure_csv_exists()
            with open(CSV_FILE, "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow([timestamp, sender, message])
        except OSError as e:
            log.error(f"Failed to write CSV row: {e}")
        try:
            ensure_excel_exists()
            wb = load_workbook(EXCEL_FILE)
            ws = wb["Messages"]
            ws.append([timestamp, sender, message])
            wb.save(EXCEL_FILE)
        except Exception as e:
            log.error(f"Failed to write Excel row: {e}")

@app.route("/webhook", methods=["GET"])
def verify_webhook():
    mode = request.args.get("hub.mode")
    token = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge")

    if mode == "subscribe" and token == VERIFY_TOKEN:
        return challenge, 200
    log.warning("Webhook verification attempt failed (bad mode/verify token).")
    return "Verification failed", 403

@app.route("/webhook", methods=["POST"])
def receive_webhook_message():
    if not is_valid_signature(request):
        log.error("Rejected webhook POST: invalid or missing signature.")
        return jsonify({"status": "invalid signature"}), 403
    data = request.get_json(silent=True) or {}
    try:
        entries = data.get("entry", [])
        for entry in entries:
            changes = entry.get("changes", [])
            for change in changes:
                value = change.get("value", {})
                messages = value.get("messages", [])
                for msg in messages:
                    try:
                        process_single_message(msg)
                    except Exception as e:
                        log.exception(f"Failed to process message {msg.get('id')}: {e}")
    except Exception as e:
        log.exception(f"Error processing webhook payload: {e}")
    return jsonify({"status": "received"}), 200

def process_single_message(msg: dict):
    sender = msg.get("from", "unknown")
    wa_timestamp = msg.get("timestamp")
    if wa_timestamp:
        try:
            timestamp = datetime.fromtimestamp(
                int(wa_timestamp), tz=timezone.utc
            ).strftime("%Y-%m-%d %H:%M:%S")
        except (ValueError, OverflowError):
            timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    else:
        timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    msg_type = msg.get("type")
    if msg_type == "text":
        message_text = msg.get("text", {}).get("body", "")
    else:
        message_text = f"[Unsupported message type: {msg_type}]"
    log.info(f"New message from {sanitize_sender(sender)}")
    log_message(timestamp, sender, message_text)

if __name__ == "__main__":
    ensure_csv_exists()
    ensure_excel_exists()
    # debug=True is convenient locally but should be False in production
    app.run(host="0.0.0.0", port=5000, debug=True)