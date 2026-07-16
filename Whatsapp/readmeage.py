"""
WhatsApp Cloud API - Incoming Message Logger
----------------------------------------------
Receives incoming WhatsApp messages via the Meta WhatsApp Cloud API webhook,
then logs (timestamp, sender_number, message) to both a CSV file and an
Excel (.xlsx) file.

Setup required on Meta side:
1. Create a Meta App -> add "WhatsApp" product.
2. In WhatsApp > Configuration, set the Webhook Callback URL to:
       https://<your-public-domain>/webhook
   (use ngrok/cloudflared during local testing, e.g. `ngrok http 5000`)
3. Set the Verify Token to the same value as VERIFY_TOKEN below.
4. Subscribe to the "messages" webhook field.

Run:
    pip install flask openpyxl
    python whatsapp_message_logger.py
"""

import os
import csv
import threading
from datetime import datetime

from flask import Flask, request, jsonify
from openpyxl import Workbook, load_workbook

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------
VERIFY_TOKEN = "your_custom_verify_token"   # must match what you set in Meta App dashboard

CSV_FILE = "whatsapp_messages.csv"
EXCEL_FILE = "whatsapp_messages.xlsx"

CSV_HEADERS = ["Timestamp", "Sender Number", "Message"]

# Lock to avoid race conditions if multiple requests hit the webhook at once
file_lock = threading.Lock()

app = Flask(__name__)


# ---------------------------------------------------------------------------
# File helpers
# ---------------------------------------------------------------------------
def ensure_csv_exists():
    if not os.path.exists(CSV_FILE):
        with open(CSV_FILE, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(CSV_HEADERS)


def ensure_excel_exists():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Messages"
        ws.append(CSV_HEADERS)
        wb.save(EXCEL_FILE)


def log_message(timestamp: str, sender: str, message: str):
    """Append a single message record to both CSV and Excel files."""
    with file_lock:
        ensure_csv_exists()
        ensure_excel_exists()

        # --- CSV ---
        with open(CSV_FILE, "a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([timestamp, sender, message])

        # --- Excel ---
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Messages"]
        ws.append([timestamp, sender, message])
        wb.save(EXCEL_FILE)


# ---------------------------------------------------------------------------
# Webhook verification (GET) - Meta calls this once when you save the webhook config
# ---------------------------------------------------------------------------
@app.route("/webhook", methods=["GET"])
def verify_webhook():
    mode = request.args.get("hub.mode")
    token = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge")

    if mode == "subscribe" and token == VERIFY_TOKEN:
        return challenge, 200
    return "Verification failed", 403


# ---------------------------------------------------------------------------
# Webhook receiver (POST) - Meta sends incoming message events here
# ---------------------------------------------------------------------------
@app.route("/webhook", methods=["POST"])
def receive_webhook():
    data = request.get_json(silent=True) or {}

    try:
        entries = data.get("entry", [])
        for entry in entries:
            changes = entry.get("changes", [])
            for change in changes:
                value = change.get("value", {})
                messages = value.get("messages", [])

                for msg in messages:
                    sender = msg.get("from", "unknown")

                    # WhatsApp gives message timestamp as unix epoch seconds
                    wa_timestamp = msg.get("timestamp")
                    if wa_timestamp:
                        timestamp = datetime.fromtimestamp(int(wa_timestamp)).strftime(
                            "%Y-%m-%d %H:%M:%S"
                        )
                    else:
                        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                    # Extract message text (handles plain text messages;
                    # extend this if you also want to handle images/buttons/etc.)
                    msg_type = msg.get("type")
                    if msg_type == "text":
                        message_text = msg["text"]["body"]
                    else:
                        message_text = f"[Unsupported message type: {msg_type}]"

                    print(f"New message from {sender}: {message_text}")
                    log_message(timestamp, sender, message_text)

    except Exception as e:
        print(f"Error processing webhook payload: {e}")

    # Always return 200 quickly so Meta doesn't retry/resend the event
    return jsonify({"status": "received"}), 200


# ---------------------------------------------------------------------------
if __name__ == "__main__":
    ensure_csv_exists()
    ensure_excel_exists()
    app.run(host="0.0.0.0", port=5000, debug=True)