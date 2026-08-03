import os
import re
import csv
import time
import hmac
import hashlib
import logging
import threading
from datetime import datetime, timezone, timedelta
from collections import defaultdict, deque
import requests
from flask import Flask, request, redirect, jsonify
from openpyxl import Workbook, load_workbook
import database.UserDB as dbimp
from urllib.parse import urlencode
from supabase import create_client, Client
import Whatsapp.new as wtpp
SUPABASE_URL = os.environ.get("SUPABASE_URL")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY")
WA_APP_ID = os.getenv("WA_APP_ID")
WA_REDIRECT_URI = os.getenv("WA_REDIRECT_URI")
GRAPH_VERSION = "v20.0"
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)
TABLE_NAME = "WhatsApp"
SCOPE = "whatsapp_business_management,whatsapp_business_messaging,business_management"
SEND_API_KEY = os.environ.get("WHATSAPP_SEND_API_KEY")
VERIFY_TOKEN = os.environ.get("WHATSAPP_VERIFY_TOKEN")
APP_SECRET = os.environ.get("WHATSAPP_APP_SECRET")
GRAPH_URL = "https://graph.facebook.com/v20.0"
REQUEST_TIMEOUT = 15
MAX_RETRIES = 3
RETRY_BACKOFF_BASE = 1.5
MAX_TEXT_LENGTH = 1800
MAX_FILE_SIZE_BYTES = { "image": 5 * 1024 * 1024, "audio": 16 * 1024 * 1024, "video": 16 * 1024 * 1024, "document": 100 * 1024 * 1024, }
VALID_MEDIA_TYPES = set(MAX_FILE_SIZE_BYTES.keys())
MAX_BUTTONS = 3
MAX_BUTTON_TITLE_LEN = 20
MAX_BUTTON_ID_LEN = 256
MAX_LIST_SECTIONS = 10
MAX_LIST_ROWS_TOTAL = 10
MAX_LIST_ROW_TITLE_LEN = 24
MAX_LIST_ROW_DESC_LEN = 72
MAX_LIST_BUTTON_TEXT_LEN = 20
CSV_FILE = r"Analytics/Report/whatsapp_messages.csv"
EXCEL_FILE = r"Analytics/Report/whatsapp_messages.xlsx"
COLUMNS_NAME = ["Timestamp", "Sender Number", "Message"]
MAX_LOG_MESSAGE_LENGTH = 4000
_request_log = defaultdict(deque)
file_lock = threading.Lock()
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
log = logging.getLogger("whatsapp_integration")
if not all([WA_APP_ID, VERIFY_TOKEN, APP_SECRET, WA_REDIRECT_URI, SEND_API_KEY]):
    log.warning("Missing one or more required WhatsApp environment variables.")
app = Flask(__name__)
app.config["MAX_CONTENT_LENGTH"] = 1 * 1024 * 1024

class InvalidPhoneNumberError(Exception):
    pass

class MessageTooLongError(Exception):
    pass

class FileTooLargeError(Exception):
    pass

def is_valid_signature(req) -> bool:
    if not APP_SECRET:
        log.error("APP_SECRET not configured — refusing to process webhook.")
        return False
    signature_header = req.headers.get("X-Hub-Signature-256", "")
    if not signature_header.startswith("sha256="):
        log.warning("Missing or malformed X-Hub-Signature-256 header.")
        return False
    received_sig = signature_header.split("sha256=", 1)[1]
    expected_sig = hmac.new(APP_SECRET.encode("utf-8"), req.get_data(), hashlib.sha256).hexdigest()
    return hmac.compare_digest(received_sig, expected_sig)

def validate_phone_number(number: str) -> str:
    if not number or not re.fullmatch(r"\+?[1-9]\d{7,14}", str(number)):
        raise InvalidPhoneNumberError(f"'{number}' is not a valid phone number.")
    return re.sub(r"[^\d+]", "", str(number))

def sanitize_field(value: str) -> str:
    if value is None:
        return ""
    value = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", str(value))
    return value.replace("\x00", "")

def sanitize_log_field(value: str) -> str:
    if value is None:
        return ""
    value = str(value)
    value = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f]", "", value)
    value = value.replace("\x00", "")
    if value[:1] in ("=", "+", "-", "@"):
        value = "'" + value
    return value[:MAX_LOG_MESSAGE_LENGTH]

def sanitize_sender(value: str) -> str:
    if not value:
        return "unknown"
    value = re.sub(r"[^0-9+]", "", str(value))
    return value[:20] or "unknown"

def validate_text_body(body: str) -> str:
    if body is None:
        body = ""
    if len(body) > MAX_TEXT_LENGTH:
        raise MessageTooLongError(f"Message is {len(body)} chars, exceeds WhatsApp's {MAX_TEXT_LENGTH}-char limit.")
    return body

def check_remote_file_size(url: str, msg_type: str):
    max_bytes = MAX_FILE_SIZE_BYTES.get(msg_type)
    if not max_bytes:
        return
    try:
        resp = requests.head(url, timeout=REQUEST_TIMEOUT, allow_redirects=True)
        content_length = resp.headers.get("Content-Length")
        if content_length is not None and int(content_length) > max_bytes:
            raise FileTooLargeError(f"File at {url} is {content_length} bytes, exceeds {max_bytes} byte limit for '{msg_type}'." )
    except requests.RequestException as e:
        log.warning(f"Could not verify remote file size for {url}: {e}")

def request_with_retry(method: str, url: str, **kwargs) -> requests.Response:
    last_exc = None
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = requests.request(method, url, timeout=REQUEST_TIMEOUT, **kwargs)
            if resp.status_code >= 500:
                raise requests.HTTPError(f"Server error {resp.status_code}")
            resp.raise_for_status()
            return resp
        except (requests.ConnectionError, requests.Timeout, requests.HTTPError) as e:
            last_exc = e
            wait = RETRY_BACKOFF_BASE ** attempt
            log.warning(f"Send attempt {attempt}/{MAX_RETRIES} failed: {e}. Retrying in {wait:.1f}s.")
            time.sleep(wait)
    raise last_exc

def require_api_key():
    if not SEND_API_KEY:
        return jsonify({"error": "Server not configured (missing API key)"}), 500
    provided = request.headers.get("X-API-Key", "")
    if provided != SEND_API_KEY:
        log.warning("Rejected /send-test call: missing/invalid X-API-Key.")
        return jsonify({"error": "Unauthorized"}), 401
    return None

def send_whatsapp_message(PHONE_NUMBER_ID, ACCESS_TOKEN, recipient_number: str, message_body: str) -> dict:
    recipient_number = validate_phone_number(recipient_number)
    message_body = validate_text_body(message_body)
    payload = { "messaging_product": "whatsapp", "recipient_type": "individual", "to": recipient_number, "type": "text", "text": {"body": message_body}, }
    url = f"{GRAPH_URL}/{PHONE_NUMBER_ID}/messages"
    resp = request_with_retry( "POST", url, headers={"Authorization": f"Bearer {ACCESS_TOKEN}", "Content-Type": "application/json"}, json=payload, )
    return resp.json()

def send_whatsapp_media(PHONE_NUMBER_ID, ACCESS_TOKEN, recipient_number: str, msg_type: str,
                         link: str, caption: str = None, filename: str = None) -> dict:
    recipient_number = validate_phone_number(recipient_number)
    if msg_type not in VALID_MEDIA_TYPES:
        raise ValueError(f"msg_type must be one of {VALID_MEDIA_TYPES}, got '{msg_type}'")
    if not link:
        raise ValueError("A 'link' URL is required to send media.")
    check_remote_file_size(link, msg_type)
    media_obj = {"link": link}
    if caption and msg_type in ("image", "video", "document"):
        media_obj["caption"] = validate_text_body(caption)
    if filename and msg_type == "document":
        media_obj["filename"] = filename
    payload = { "messaging_product": "whatsapp", "recipient_type": "individual", "to": recipient_number, "type": msg_type, msg_type: media_obj, }
    url = f"{GRAPH_URL}/{PHONE_NUMBER_ID}/messages"
    resp = request_with_retry( "POST", url, headers={"Authorization": f"Bearer {ACCESS_TOKEN}", "Content-Type": "application/json"}, json=payload, )
    return resp.json()

def send_whatsapp_location(PHONE_NUMBER_ID, ACCESS_TOKEN, recipient_number: str, latitude: float, longitude: float, name: str = None, address: str = None) -> dict:
    recipient_number = validate_phone_number(recipient_number)
    try:
        lat = float(latitude)
        lng = float(longitude)
    except (TypeError, ValueError):
        raise ValueError("latitude/longitude must be numbers.")
    if not (-90 <= lat <= 90) or not (-180 <= lng <= 180):
        raise ValueError("latitude must be in [-90, 90] and longitude in [-180, 180].")
    location_obj = {"latitude": lat, "longitude": lng}
    if name:
        location_obj["name"] = sanitize_field(name)[:1000]
    if address:
        location_obj["address"] = sanitize_field(address)[:1000]
    payload = { "messaging_product": "whatsapp","recipient_type": "individual","to": recipient_number, "type": "location","location": location_obj, }
    url = f"{GRAPH_URL}/{PHONE_NUMBER_ID}/messages"
    resp = request_with_retry( "POST", url, headers={"Authorization": f"Bearer {ACCESS_TOKEN}", "Content-Type": "application/json"}, json=payload,)
    return resp.json()

def send_whatsapp_reply_buttons(PHONE_NUMBER_ID, ACCESS_TOKEN, recipient_number: str,body_text: str, buttons: list) -> dict:
    recipient_number = validate_phone_number(recipient_number)
    body_text = validate_text_body(body_text)
    if not buttons or len(buttons) > MAX_BUTTONS:
        raise ValueError(f"Provide 1-{MAX_BUTTONS} buttons, got {len(buttons) if buttons else 0}.")
    formatted_buttons = []
    seen_ids = set()
    for b in buttons:
        btn_id = str(b.get("id", "")).strip()
        title = str(b.get("title", "")).strip()
        if not btn_id or not title:
            raise ValueError("Each button needs a non-empty 'id' and 'title'.")
        if len(btn_id) > MAX_BUTTON_ID_LEN:
            raise ValueError(f"Button id exceeds {MAX_BUTTON_ID_LEN} chars.")
        if len(title) > MAX_BUTTON_TITLE_LEN:
            raise ValueError(f"Button title '{title}' exceeds WhatsApp's {MAX_BUTTON_TITLE_LEN}-char limit.")
        if btn_id in seen_ids:
            raise ValueError(f"Duplicate button id '{btn_id}'.")
        seen_ids.add(btn_id)
        formatted_buttons.append({"type": "reply", "reply": {"id": btn_id, "title": title}})
    payload = { "messaging_product": "whatsapp", "recipient_type": "individual", "to": recipient_number, "type": "interactive", "interactive": { "type": "button", "body": {"text": body_text}, "action": {"buttons": formatted_buttons}, },}
    url = f"{GRAPH_URL}/{PHONE_NUMBER_ID}/messages"
    resp = request_with_retry( "POST", url, headers={"Authorization": f"Bearer {ACCESS_TOKEN}", "Content-Type": "application/json"}, json=payload, )
    return resp.json()

def send_whatsapp_list(PHONE_NUMBER_ID, ACCESS_TOKEN, recipient_number: str, body_text: str, button_text: str, sections: list) -> dict:
    recipient_number = validate_phone_number(recipient_number)
    body_text = validate_text_body(body_text)
    if not button_text or len(button_text) > MAX_LIST_BUTTON_TEXT_LEN:
        raise ValueError(f"'button_text' must be 1-{MAX_LIST_BUTTON_TEXT_LEN} chars.")
    if not sections or len(sections) > MAX_LIST_SECTIONS:
        raise ValueError(f"Provide 1-{MAX_LIST_SECTIONS} sections.")
    total_rows = 0
    formatted_sections = []
    seen_row_ids = set()
    for section in sections:
        title = str(section.get("title", "")).strip()
        rows = section.get("rows", [])
        if not rows:
            raise ValueError(f"Section '{title}' has no rows.")
        formatted_rows = []
        for row in rows:
            row_id = str(row.get("id", "")).strip()
            row_title = str(row.get("title", "")).strip()
            row_desc = str(row.get("description", "")).strip()
            if not row_id or not row_title:
                raise ValueError("Each row needs a non-empty 'id' and 'title'.")
            if len(row_title) > MAX_LIST_ROW_TITLE_LEN:
                raise ValueError(f"Row title '{row_title}' exceeds {MAX_LIST_ROW_TITLE_LEN}-char limit.")
            if len(row_desc) > MAX_LIST_ROW_DESC_LEN:
                raise ValueError(f"Row description exceeds {MAX_LIST_ROW_DESC_LEN}-char limit.")
            if row_id in seen_row_ids:
                raise ValueError(f"Duplicate row id '{row_id}'.")
            seen_row_ids.add(row_id)
            row_obj = {"id": row_id, "title": row_title}
            if row_desc:
                row_obj["description"] = row_desc
            formatted_rows.append(row_obj)
            total_rows += 1
        if total_rows > MAX_LIST_ROWS_TOTAL:
            raise ValueError(f"Total rows across all sections exceeds {MAX_LIST_ROWS_TOTAL}.")
        formatted_sections.append({"title": title, "rows": formatted_rows})
    payload = { "messaging_product": "whatsapp", "recipient_type": "individual", "to": recipient_number, "type": "interactive", "interactive": { "type": "list", "body": {"text": body_text}, "action": {"button": button_text, "sections": formatted_sections}, }, }
    url = f"{GRAPH_URL}/{PHONE_NUMBER_ID}/messages"
    resp = request_with_retry("POST", url, headers={"Authorization": f"Bearer {ACCESS_TOKEN}", "Content-Type": "application/json"}, json=payload,)
    return resp.json()

def ensure_csv_exists():
    os.makedirs(os.path.dirname(CSV_FILE), exist_ok=True)
    if not os.path.exists(CSV_FILE):
        with open(CSV_FILE, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow(COLUMNS_NAME)

def ensure_excel_exists():
    os.makedirs(os.path.dirname(EXCEL_FILE), exist_ok=True)
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Messages"
        ws.append(COLUMNS_NAME)
        wb.save(EXCEL_FILE)

def log_message(timestamp: str, sender: str, message: str):
    sender = sanitize_sender(sender)
    message = sanitize_log_field(message)
    with file_lock:
        try:
            ensure_csv_exists()
            with open(CSV_FILE, "a", newline="", encoding="utf-8") as f:
                writer = csv.writer(f)
                writer.writerow([timestamp, sender, message])
        except OSError as e:
            log.error(f"Failed to write CSV row: {e}")
        try:
            ensure_excel_exists()
            wb = load_workbook(EXCEL_FILE)
            ws = wb["Messages"]
            ws.append([timestamp, sender, message])
            wb.save(EXCEL_FILE)
        except Exception as e:
            log.error(f"Failed to write Excel row: {e}")

def get_user_for_phone_number_id(incoming_id: str):
    if not incoming_id:
        return None
    rows = dbimp.select_rows(TABLE_NAME, select="id,Access_token", filters={"Account_id": incoming_id})
    return rows[0] if rows else None

def process_single_message(msg: dict):
    sender = msg.get("from", "unknown")
    wa_timestamp = msg.get("timestamp")
    if wa_timestamp:
        try:
            timestamp = datetime.fromtimestamp(int(wa_timestamp), tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
        except (ValueError, OverflowError):
            timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    else:
        timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    msg_type = msg.get("type")
    if msg_type == "text":
        message_text = msg.get("text", {}).get("body", "")
    else:
        message_text = f"[Unsupported message type: {msg_type}]"
    log.info(f"New message from {sanitize_sender(sender)}")
    log_message(timestamp, sender, message_text)

def check_user_id(user_id):
    exist = dbimp.select_rows(TABLE_NAME, select="id", filters={"id": user_id})
    return bool(exist)

def refresh_token(user_id, access_token):
    resp = requests.get( f"https://graph.facebook.com/{GRAPH_VERSION}/oauth/access_token", params={ "grant_type": "fb_exchange_token", "client_id": WA_APP_ID, "client_secret": APP_SECRET, "fb_exchange_token": access_token, }, ).json()
    new_token = resp.get("access_token")
    seconds = resp.get("expires_in")
    if new_token and seconds:
        new_expire = datetime.now(timezone.utc) + timedelta(seconds=seconds)
        dbimp.update_rows( TABLE_NAME, {"Access_token": new_token, "Token_expire": new_expire.isoformat()}, filters={"id": user_id}, )
        return new_token
    return access_token

@app.route("/auth/whatsapp/login")
def whatsapp_login():
    user_id = request.args.get("user_id")  # /auth/whatsapp/login?user_id=<some_id>
    if not user_id:
        return jsonify({"error": "user_id is required"}), 400
    params = { "client_id": WA_APP_ID, "redirect_uri": WA_REDIRECT_URI, "scope": SCOPE, "response_type": "code", "state": user_id, }
    auth_url = f"https://www.facebook.com/{GRAPH_VERSION}/dialog/oauth?" + urlencode(params)
    return redirect(auth_url)

@app.route("/auth/whatsapp/callback")
def whatsapp_callback():
    code = request.args.get("code")
    user_id = request.args.get("state")
    if not code:
        return jsonify({"error": "missing code"}), 400
    if not user_id:
        return jsonify({"error": "missing user id"}), 400
    if not check_user_id(user_id):
        return jsonify({"error": "invalid user id"}), 400
    token_resp = requests.get( f"https://graph.facebook.com/{GRAPH_VERSION}/oauth/access_token", params={ "client_id": WA_APP_ID, "client_secret": APP_SECRET, "redirect_uri": WA_REDIRECT_URI, "code": code, },).json()
    short_token = token_resp.get("access_token")
    if not short_token:
        return jsonify({"error": "token exchange failed", "details": token_resp}), 400
    long_resp = requests.get(f"https://graph.facebook.com/{GRAPH_VERSION}/oauth/access_token", params={ "grant_type": "fb_exchange_token", "client_id": WA_APP_ID, "client_secret": APP_SECRET, "fb_exchange_token": short_token, },).json()
    long_token = long_resp.get("access_token")
    seconds = long_resp.get("expires_in")
    if not long_token or not seconds:
        return jsonify({"error": "token exchange failed", "details": long_resp}), 400
    expire_time = datetime.now(timezone.utc) + timedelta(seconds=seconds)
    debug = requests.get(f"https://graph.facebook.com/{GRAPH_VERSION}/debug_token", params={"input_token": long_token, "access_token": f"{WA_APP_ID}|{APP_SECRET}"}, ).json()
    granular_scopes = debug.get("data", {}).get("granular_scopes", [])
    waba_ids = []
    for scope in granular_scopes:
        if scope.get("scope") == "whatsapp_business_management":
            waba_ids.extend(scope.get("target_ids", []))
    waba_id = waba_ids[0] if waba_ids else None
    phone_number_id = None
    display_number = None
    if waba_id:
        phones = requests.get( f"https://graph.facebook.com/{GRAPH_VERSION}/{waba_id}/phone_numbers", params={"access_token": long_token}, ).json()
        numbers = phones.get("data", [])
        if numbers:
            phone_number_id = numbers[0].get("id")
            display_number = numbers[0].get("display_phone_number")
    try:
        dbimp.update_rows(TABLE_NAME, {"Access_token": long_token,"Timestamp": datetime.now(timezone.utc).isoformat(),"Token_expire": expire_time.isoformat(),"Bussiness_id": waba_id, "Account_id": phone_number_id, "Phone_no": display_number,}, filters={"id": user_id},)
    except Exception as e:
        return jsonify({"error": "token stored failed to save", "details": str(e)}), 500
    return jsonify({"user_id": user_id, "waba_id": waba_id,"phone_number_id": phone_number_id, "access_token": long_token, })

@app.route("/webhook", methods=["GET"])
def verify_webhook():
    mode = request.args.get("hub.mode")
    token = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge")
    if mode == "subscribe" and token == VERIFY_TOKEN:
        return challenge, 200
    log.warning("Webhook verification attempt failed (bad mode/verify token).")
    return "Verification failed", 403

@app.route("/webhook", methods=["POST"])
def receive_webhook_message():
    if not is_valid_signature(request):
        log.error("Rejected webhook POST: invalid or missing signature.")
        return jsonify({"status": "invalid signature"}), 403
    data = request.get_json(silent=True) or {}
    try:
        entries = data.get("entry", [])
        for entry in entries:
            changes = entry.get("changes", [])
            for change in changes:
                value = change.get("value", {})
                incoming_id = value.get("metadata", {}).get("phone_number_id")
                user_row = get_user_for_phone_number_id(incoming_id)
                if not user_row:
                    log.warning(f"Rejected webhook payload: unrecognized phone_number_id ({incoming_id!r}).")
                    continue
                messages = value.get("messages", [])
                for msg in messages:
                    try:
                        process_single_message(msg)
                    except Exception as e:
                        log.exception(f"Failed to process message {msg.get('id')}: {e}")
    except Exception as e:
        log.exception(f"Error processing webhook payload: {e}")
    return jsonify({"status": "received"}), 200

@app.route("/send-test", methods=["POST"])
def test_send():
    auth_error = require_api_key()
    if auth_error:
        return auth_error
    user_id = request.args.get("user_id")
    account_id = request.args.get("account_id")
    if not user_id or not account_id:
        return jsonify({"error": "'user_id' and 'account_id' are required"}), 400
    phone_number_raw = request.args.get("phone_number")
    if not phone_number_raw:
        return jsonify({"error": "'phone_number' query param is required"}), 400
    try:
        phoneno = int(phone_number_raw)
    except ValueError:
        return jsonify({"error": "'phone_number' must be numeric"}), 400
    rows = dbimp.select_rows( TABLE_NAME, select="Access_token,Token_expire", filters={"id": user_id, "Account_id": account_id},)
    if not rows:
        return jsonify({"error": "no whatsapp account linked"}), 404
    row = rows[0]
    access_token = row["Access_token"]
    token_expiry = row["Token_expire"]
    if not access_token or not token_expiry:
        return jsonify({"error": "missing access_token"}), 400
    token_expiry = datetime.fromisoformat(token_expiry)
    if token_expiry.tzinfo is None:
        token_expiry = token_expiry.replace(tzinfo=timezone.utc)
    if token_expiry - datetime.now(timezone.utc) < timedelta(days=2):
        access_token = refresh_token(user_id, access_token)
    data = request.get_json(silent=True)
    if not data or "phone" not in data:
        return jsonify({"error": "Please provide at least 'phone'"}), 400
    phone = data["phone"]
    msg_type = data.get("type", "text")
    try:
        if msg_type == "text":
            if "msg" not in data:
                return jsonify({"error": "'msg' is required for type 'text'"}), 400
            result = send_whatsapp_message(phoneno, access_token, phone, data["msg"])
        elif msg_type in VALID_MEDIA_TYPES:
            if "link" not in data:
                return jsonify({"error": f"'link' is required for type '{msg_type}'"}), 400
            result = send_whatsapp_media( phoneno, access_token, phone, msg_type, link=data["link"], caption=data.get("caption"), filename=data.get("filename"),)
        elif msg_type == "location":
            if "latitude" not in data or "longitude" not in data:
                return jsonify({"error": "'latitude' and 'longitude' are required"}), 400
            result = send_whatsapp_location(phoneno, access_token, phone, data["latitude"], data["longitude"], name=data.get("name"), address=data.get("address"),)
        elif msg_type == "button":
            if "body" not in data or "buttons" not in data:
                return jsonify({"error": "'body' and 'buttons' are required"}), 400
            result = send_whatsapp_reply_buttons(phoneno, access_token, phone, data["body"], data["buttons"])
        elif msg_type == "list":
            if "body" not in data or "button_text" not in data or "sections" not in data:
                return jsonify({"error": "'body', 'button_text', and 'sections' are required"}), 400
            result = send_whatsapp_list( phoneno, access_token, phone, data["body"], data["button_text"], data["sections"] )
        else:
            return jsonify({"error": f"Unsupported type '{msg_type}'"}), 400
        return jsonify({"status": "sent", "details": result}), 200
    except InvalidPhoneNumberError as e:
        return jsonify({"error": str(e)}), 400
    except MessageTooLongError as e:
        return jsonify({"error": str(e)}), 400
    except FileTooLargeError as e:
        return jsonify({"error": str(e)}), 413
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    except requests.RequestException as e:
        log.error(f"Send failed after retries: {e}")
        return jsonify({"error": "Failed to send message after retries"}), 502
    except Exception as e:
        log.exception(f"Unexpected error in /send-test: {e}")
        return jsonify({"error": "Internal error"}), 500

if __name__ == "__main__":
    ensure_csv_exists()
    ensure_excel_exists()
    # debug=False in production — the Werkzeug debugger is an RCE risk if exposed.
    app.run(host="0.0.0.0", port=5000, debug=False)